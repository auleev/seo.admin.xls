import requests
import openpyxl
from requests.exceptions import SSLError
from urllib.parse import urlparse
from openpyxl.styles import PatternFill

# Fonction pour obtenir l'URL de recette en fonction de l'URL de production
def get_recette_url(prod_url):
    if 'promocroisiere.com' in prod_url:
        return prod_url.replace('https://www.promocroisiere.com', 'https://p5-www.promocroisiere.com')
    elif 'promovacances.com' in prod_url:
        return prod_url.replace('https://www.promovacances.com', 'https://p6-www.promovacances.com')
    else:
        return None

# Fonction pour obtenir les détails HTTP d'une URL
def get_http_details(url):
    try:
        response = requests.get(url, verify=False)  # Désactiver la vérification SSL
        http_code = response.status_code
        redirection_chain = ' > '.join([str(resp.status_code) for resp in response.history] + [str(http_code)])
        # Extraire uniquement l'URI (chemin relatif) de l'URL finale
        final_url = urlparse(response.url).path
        return redirection_chain, len(response.history), final_url
    except SSLError as ssl_err:
        return 'Erreur SSL', 'N/A', 'N/A'
    except requests.exceptions.RequestException as e:
        return 'Erreur', 'N/A', 'N/A'

# Charger le fichier Excel
workbook = openpyxl.load_workbook('urls.xlsx')
worksheet = workbook['Feuil1']  # Remplacez 'Feuil1' par le nom de votre feuille

# Insérer une nouvelle colonne pour "URL CIBLE" après la première colonne
worksheet.cell(1, 1, 'URL Production')
worksheet.insert_cols(2)
worksheet.cell(1, 2, 'URL Recette')  # Ajouter l'en-tête de la nouvelle colonne

# Ajouter des en-têtes pour les nouvelles colonnes
worksheet.cell(1, 3, 'Status HTTP (Prod)')
worksheet.cell(1, 4, 'Nombre de Redirections (Prod)')
worksheet.cell(1, 5, 'URL Finale (Prod)')
worksheet.cell(1, 6, 'Status HTTP (Recette)')
worksheet.cell(1, 7, 'Nombre de Redirections (Recette)')
worksheet.cell(1, 8, 'URL Finale (Recette)')
worksheet.cell(1, 9, 'Écart Status HTTP')
worksheet.cell(1, 10, 'Écart Nombre de Redirections')
worksheet.cell(1, 11, 'Écart URL Finale')

# Définir la mise en forme pour les cellules avec des écarts
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
# Boucle sur chaque ligne du fichier Excel
for row in range(2, worksheet.max_row + 1):  # Commence à la ligne 2 pour éviter l'en-tête
    prod_url = worksheet.cell(row, 1).value

    # Obtenir l'URL de recette
    recette_url = get_recette_url(prod_url)
    worksheet.cell(row, 2).value = recette_url if recette_url else 'N/A'

    # Obtenir les détails pour l'URL de production
    prod_redirection_chain, prod_redirections_count, prod_final_url = get_http_details(prod_url)

    # Écrire les données pour l'URL de production
    worksheet.cell(row, 3).value = prod_redirection_chain
    worksheet.cell(row, 4).value = prod_redirections_count
    worksheet.cell(row, 5).value = prod_final_url

    # Obtenir les détails pour l'URL de recette
    if recette_url:
        recette_redirection_chain, recette_redirections_count, recette_final_url = get_http_details(recette_url)

        # Écrire les données pour l'URL de recette
        worksheet.cell(row, 6).value = recette_redirection_chain
        worksheet.cell(row, 7).value = recette_redirections_count
        worksheet.cell(row, 8).value = recette_final_url

        # Calculer les écarts
        status_http_ecart = 'Identique' if prod_final_url == recette_final_url else 'Différent'
        redirections_ecart = 'Identique' if prod_redirections_count == recette_redirections_count else 'Différent'
        url_finale_ecart = 'Identique' if prod_redirection_chain == recette_redirection_chain else 'Différent'

        # Écrire les écarts
        worksheet.cell(row, 9).value = status_http_ecart
        worksheet.cell(row, 10).value = redirections_ecart
        worksheet.cell(row, 11).value = url_finale_ecart

        # Appliquer la mise en forme rouge si l'écart est différent
        if status_http_ecart == 'Différent':
            worksheet.cell(row, 9).fill = red_fill
        if redirections_ecart == 'Différent':
            worksheet.cell(row, 10).fill = red_fill
        if url_finale_ecart == 'Différent':
            worksheet.cell(row, 11).fill = red_fill
    else:
        # Si l'URL de recette n'est pas trouvée
        worksheet.cell(row, 6).value = 'N/A'
        worksheet.cell(row, 7).value = 'N/A'
        worksheet.cell(row, 8).value = 'N/A'
        worksheet.cell(row, 9).value = 'N/A'
        worksheet.cell(row, 10).value = 'N/A'
        worksheet.cell(row, 11).value = 'N/A'

# Enregistrer les modifications dans le fichier Excel
workbook.save('urls_comparaison.xlsx')
